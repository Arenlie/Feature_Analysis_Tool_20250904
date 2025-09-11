import pandas as pd
import openpyxl
from PyQt6.QtWidgets import QMessageBox
from openpyxl.styles import Alignment


def dataToDWTable(inputFile, outputFile):
    data_all = pd.read_excel(inputFile, sheet_name='输入参数')

    data_2700 = pd.DataFrame(
        columns=["边缘控制器编号", "IP地址", "主机MAC", "主机序列号", "板卡编号", "板卡出厂编号", "板卡类型",
                 "板卡是否启用",
                 "通道编号", "测点（通道）类型", "设备名称", "测点（点位）名称", "键相类型", "工作转速", "电机额定转速",
                 "电机同步转速", "电源频率", "电机转子条数", "轴承型号", "轴承生产厂家", "齿轮齿数Z", "叶轮叶片数目",
                 "导叶叶片数目"]
    )

    # 初始化空列表存储行
    rows = []

    # 生成完整的板卡编号（C01-C08）和通道编号（CH01-CH04）
    all_cards_1 = [f"C{str(i).zfill(2)}" for i in range(1, 9)]  # C01-C08
    all_cards_2 = [f"C{str(i).zfill(2)}" for i in range(1, 3)]  # C01-C02
    all_channels_1 = [f"CH0{j}" for j in range(1, 5)]  # CH01-CH04
    all_channels_2 = [f"CH0{j}" for j in range(1, 9)]  # CH01-CH08

    # 初始化一个字典，用来存储每个主机MAC的已有板卡和通道信息
    existing_data = {}

    # 第一步：遍历每一行数据，保留原始数据并记录已有的板卡和通道信息
    for index, row in data_all.iterrows():
        if str(row["通道编码"])[:6] == "50294D" or str(row["通道编码"])[:6] == "50293D":
            # 获取主机MAC
            host_mac = row["通道编码"][:-3]

            # 提取已有的板卡编号和通道编号
            card_number = "C" + row["通道编码"][-3:-1]
            channel_number = "CH0" + row["通道编码"][-1]

            card_type = "低速卡" if row["传感器类型"] in ["温度"] else "高速卡"

            if row["传感器类型"] in ["加速度", "温度", "转速"]:
                point_type = row["传感器类型"]
            else:
                raise ValueError("不支持加速度、温度、转速以外的传感器，请检查并删除")

            # 将原始数据加入到结果中
            row_data = {
                "边缘控制器编号": "", "IP地址": "", "主机序列号": "", "板卡出厂编号": "",
                "主机MAC": host_mac, "板卡编号": card_number,
                # "测点（通道）类型": "动态电压" if row["传感器类型"] in ["电流谱", "电压谱"] else row["传感器类型"], (2025-09-09废弃修改)
                "通道编号": channel_number, "测点（通道）类型": point_type,
                "设备名称": row["设备名称"], "测点（点位）名称": row["测点名称"], "工作转速": row["工作转速"],
                "电机额定转速": row["电机额定转速"], "电机同步转速": row["电机同步转速"],
                "电源频率": row["电源频率"], "电机转子条数": row["电机转子条数"], "轴承型号": row["轴承型号"],
                "轴承生产厂家": row["轴承生产厂家"], "齿轮齿数Z": row["齿轮齿数Z"],
                "叶轮叶片数目": row["叶轮叶片数目"], "导叶叶片数目": row["导叶叶片数目"],
                "板卡类型": card_type,
                "板卡是否启用": "是" if row["测点名称"] else "否",
                "键相类型": "" if pd.isna(row["工作转速"])
                else "外部键相" if len(str(row["工作转速"])) == 7
                else "虚拟键相" if 2 <= len(str(row["工作转速"])) < 7 else None
            }
            rows.append(row_data)

            # 如果该主机MAC还没有记录，初始化
            # if host_mac not in existing_data:
            #     existing_data[host_mac] = {}
            # 如果该主机MAC还没有记录，初始化
            if host_mac not in existing_data:
                existing_data[host_mac] = {
                    "tags": [row["网关型号"]]  # 初始化标签列表
                }

            # 如果该板卡还没有记录，初始化
            if card_number not in existing_data[host_mac]:
                existing_data[host_mac][card_number] = []

            # 记录已有的通道编号
            existing_data[host_mac][card_number].append(channel_number)
    # 第二步：补充缺失的板卡和通道，并保证高速卡和低速卡的数量均衡
    for host_mac, cards in existing_data.items():
        # 记录当前主机MAC的板卡计数，确保高速卡和低速卡均衡
        high_speed_count = 0
        low_speed_count = 0
        card_keys = list(cards.keys())
        # 读取当前主机MAC的标签
        tags = cards.get("tags", [])
        if tags[0] == "DW2700":
            for card in all_cards_1:
                # 确定板卡类型：前4个为高速卡，后4个为低速卡
                if high_speed_count < 4:
                    card_type = "高速卡"
                    high_speed_count += 1
                else:
                    card_type = "低速卡"
                    low_speed_count += 1

                # 如果该板卡已经存在，则检查其通道并补充缺失的通道
                if card in card_keys:
                    if card != "tags":  # 跳过标签
                        existing_channels = cards[card]
                        # 补充该板卡缺失的通道
                        for channel in all_channels_1:
                            if channel not in existing_channels:
                                if card_type == "高速卡":
                                    row_data = {
                                        "边缘控制器编号": "", "IP地址": "", "主机序列号": "", "板卡出厂编号": "",
                                        "主机MAC": host_mac, "板卡编号": card,
                                        "通道编号": channel, "测点（通道）类型": "加速度",  # 初始化为空
                                        "设备名称": "无", "测点（点位）名称": "空通道",  # 初始化为空
                                        "工作转速": "", "电机额定转速": "", "电机同步转速": "",
                                        "电源频率": "", "电机转子条数": "", "轴承型号": "",
                                        "轴承生产厂家": "", "齿轮齿数Z": "", "叶轮叶片数目": "",
                                        "导叶叶片数目": "", "板卡类型": card_type,
                                        "板卡是否启用": "否", "键相类型": None
                                    }
                                elif card_type == "低速卡":
                                    row_data = {
                                        "边缘控制器编号": "", "IP地址": "", "主机序列号": "", "板卡出厂编号": "",
                                        "主机MAC": host_mac, "板卡编号": card,
                                        "通道编号": channel, "测点（通道）类型": "普通电压",  # 初始化为空
                                        "设备名称": "无", "测点（点位）名称": "空通道",  # 初始化为空
                                        "工作转速": "", "电机额定转速": "", "电机同步转速": "",
                                        "电源频率": "", "电机转子条数": "", "轴承型号": "",
                                        "轴承生产厂家": "", "齿轮齿数Z": "", "叶轮叶片数目": "",
                                        "导叶叶片数目": "", "板卡类型": card_type,
                                        "板卡是否启用": "否", "键相类型": None
                                    }
                                rows.append(row_data)

                # 如果该板卡不存在，则为该板卡生成4个通道的数据
                else:
                    for channel in all_channels_1:
                        if card_type == "高速卡":
                            row_data = {
                                "边缘控制器编号": "", "IP地址": "", "主机序列号": "", "板卡出厂编号": "",
                                "主机MAC": host_mac, "板卡编号": card,
                                "通道编号": channel, "测点（通道）类型": "加速度",  # 初始化为空
                                "设备名称": "无", "测点（点位）名称": "空通道",  # 初始化为空
                                "工作转速": "", "电机额定转速": "", "电机同步转速": "",
                                "电源频率": "", "电机转子条数": "", "轴承型号": "",
                                "轴承生产厂家": "", "齿轮齿数Z": "", "叶轮叶片数目": "",
                                "导叶叶片数目": "", "板卡类型": card_type,
                                "板卡是否启用": "否", "键相类型": None
                            }
                        elif card_type == "低速卡":
                            row_data = {
                                "边缘控制器编号": "", "IP地址": "", "主机序列号": "", "板卡出厂编号": "",
                                "主机MAC": host_mac, "板卡编号": card,
                                "通道编号": channel, "测点（通道）类型": "普通电压",  # 初始化为空
                                "设备名称": "无", "测点（点位）名称": "空通道",  # 初始化为空
                                "工作转速": "", "电机额定转速": "", "电机同步转速": "",
                                "电源频率": "", "电机转子条数": "", "轴承型号": "",
                                "轴承生产厂家": "", "齿轮齿数Z": "", "叶轮叶片数目": "",
                                "导叶叶片数目": "", "板卡类型": card_type,
                                "板卡是否启用": "否", "键相类型": None
                            }
                        rows.append(row_data)
        elif tags[0] == "DW2300":
            for card in all_cards_2:
                # 确定板卡类型：前4个为高速卡，后4个为低速卡
                if high_speed_count < 1:
                    card_type = "高速卡"
                    high_speed_count += 1
                else:
                    card_type = "低速卡"
                    low_speed_count += 1

                # 如果该板卡已经存在，则检查其通道并补充缺失的通道
                if card in card_keys:
                    if card != "tags":  # 跳过标签
                        existing_channels = cards[card]
                        # 补充该板卡缺失的通道
                        for channel in all_channels_2:
                            if channel not in existing_channels:
                                row_data = {
                                    "边缘控制器编号": "", "IP地址": "", "主机序列号": "", "板卡出厂编号": "",
                                    "主机MAC": host_mac, "板卡编号": card,
                                    "通道编号": channel, "测点（通道）类型": "普通电压",  # 初始化为空
                                    "设备名称": "无", "测点（点位）名称": "空通道",  # 初始化为空
                                    "工作转速": "", "电机额定转速": "", "电机同步转速": "",
                                    "电源频率": "", "电机转子条数": "", "轴承型号": "",
                                    "轴承生产厂家": "", "齿轮齿数Z": "", "叶轮叶片数目": "",
                                    "导叶叶片数目": "", "板卡类型": card_type,
                                    "板卡是否启用": "否", "键相类型": None
                                }
                                rows.append(row_data)

                # 如果该板卡不存在，则为该板卡生成4个通道的数据
                else:
                    for channel in all_channels_2:
                        row_data = {
                            "边缘控制器编号": "", "IP地址": "", "主机序列号": "", "板卡出厂编号": "",
                            "主机MAC": host_mac, "板卡编号": card,
                            "通道编号": channel, "测点（通道）类型": "普通电压",  # 初始化为空
                            "设备名称": "无", "测点（点位）名称": "空通道",  # 初始化为空
                            "工作转速": "", "电机额定转速": "", "电机同步转速": "",
                            "电源频率": "", "电机转子条数": "", "轴承型号": "",
                            "轴承生产厂家": "", "齿轮齿数Z": "", "叶轮叶片数目": "",
                            "导叶叶片数目": "", "板卡类型": card_type,
                            "板卡是否启用": "否", "键相类型": None
                        }
                        rows.append(row_data)

    # 最终rows列表中将包括每个主机MAC的完整32行数据（8个板卡，每个板卡4个通道）

    data_2700 = pd.DataFrame(rows, columns=data_2700.columns)
    data_2700 = data_2700.sort_values(by=["主机MAC", "板卡编号", "通道编号"])

    with pd.ExcelWriter(outputFile, engine='openpyxl') as writer:
        for mac, group in data_2700.groupby('主机MAC'):
            # print(f"Current sheet name: {str(mac)}")
            group.to_excel(writer, index=False, sheet_name=f"{mac}")

    workbook = openpyxl.load_workbook(outputFile)

    for sheet in workbook.worksheets:
        if sheet.title.startswith("50294D") or sheet.title.startswith("50293D"):
            # 合并第前四列中相同的数值的单元格
            merge_cells_in_column_1(sheet, column_index=3)
            # 合并第五-八列中相同的数值的单元格
            merge_cells_in_column_2(sheet, column_index=5)

    workbook.save(outputFile)


def merge_cells_in_column_1(sheet, column_index):
    start_row = 2  # 从第2行开始（第1行是表头）
    end_row = sheet.max_row

    current_value = None
    merge_start = None

    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row=row, column=column_index)
        if cell.value != current_value:
            if merge_start and row - 1 > merge_start:
                sheet.merge_cells(start_row=merge_start, start_column=column_index + 1,
                                  end_row=row - 1, end_column=column_index + 1)
                align_merged_cells(sheet, merge_start, row - 1, column_index + 1)
                sheet.merge_cells(start_row=merge_start, start_column=column_index,
                                  end_row=row - 1, end_column=column_index)
                align_merged_cells(sheet, merge_start, row - 1, column_index)
                sheet.merge_cells(start_row=merge_start, start_column=column_index - 1,
                                  end_row=row - 1, end_column=column_index - 1)
                align_merged_cells(sheet, merge_start, row - 1, column_index - 1)
                sheet.merge_cells(start_row=merge_start, start_column=column_index - 2,
                                  end_row=row - 1, end_column=column_index - 2)
                align_merged_cells(sheet, merge_start, row - 1, column_index - 2)
            current_value = cell.value
            merge_start = row
        elif row == end_row:  # 如果到达最后一行
            sheet.merge_cells(start_row=merge_start, start_column=column_index + 1,
                              end_row=row, end_column=column_index + 1)
            align_merged_cells(sheet, merge_start, row, column_index + 1)
            sheet.merge_cells(start_row=merge_start, start_column=column_index,
                              end_row=row, end_column=column_index)
            align_merged_cells(sheet, merge_start, row, column_index)
            sheet.merge_cells(start_row=merge_start, start_column=column_index - 1,
                              end_row=row, end_column=column_index - 1)
            align_merged_cells(sheet, merge_start, row, column_index - 1)
            sheet.merge_cells(start_row=merge_start, start_column=column_index - 2,
                              end_row=row, end_column=column_index - 2)
            align_merged_cells(sheet, merge_start, row, column_index - 2)


def merge_cells_in_column_2(sheet, column_index):
    start_row = 2  # 从第2行开始（第1行是表头）
    end_row = sheet.max_row

    current_value = None
    merge_start = None

    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row=row, column=column_index)
        if cell.value != current_value:
            if merge_start and row - 1 > merge_start:
                sheet.merge_cells(start_row=merge_start, start_column=column_index,
                                  end_row=row - 1, end_column=column_index)
                align_merged_cells(sheet, merge_start, row - 1, column_index)
                sheet.merge_cells(start_row=merge_start, start_column=column_index + 1,
                                  end_row=row - 1, end_column=column_index + 1)
                align_merged_cells(sheet, merge_start, row - 1, column_index + 1)
                sheet.merge_cells(start_row=merge_start, start_column=column_index + 2,
                                  end_row=row - 1, end_column=column_index + 2)
                align_merged_cells(sheet, merge_start, row - 1, column_index + 2)
                sheet.merge_cells(start_row=merge_start, start_column=column_index + 3,
                                  end_row=row - 1, end_column=column_index + 3)
                align_merged_cells(sheet, merge_start, row - 1, column_index + 3)
            current_value = cell.value
            merge_start = row
        elif row == end_row:  # 如果到达最后一行
            sheet.merge_cells(start_row=merge_start, start_column=column_index,
                              end_row=row, end_column=column_index)
            align_merged_cells(sheet, merge_start, row, column_index)
            sheet.merge_cells(start_row=merge_start, start_column=column_index + 1,
                              end_row=row, end_column=column_index + 1)
            align_merged_cells(sheet, merge_start, row, column_index + 1)
            sheet.merge_cells(start_row=merge_start, start_column=column_index + 2,
                              end_row=row, end_column=column_index + 2)
            align_merged_cells(sheet, merge_start, row, column_index + 2)
            sheet.merge_cells(start_row=merge_start, start_column=column_index + 3,
                              end_row=row, end_column=column_index + 3)
            align_merged_cells(sheet, merge_start, row, column_index + 3)


def align_merged_cells(sheet, start_row, end_row, column_index):
    """对齐合并后的单元格"""
    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row=row, column=column_index)
        cell.alignment = Alignment(horizontal='center', vertical='center')


if __name__ == "__main__":
    dataToDWTable(r"D:\项目资料\特征解析工具汇编\PHM2.0打包\data_all - 2300测试.xlsx",
                    r"D:\项目资料\特征解析工具汇编\PHM2.0打包\data_all - 2700导入表.xlsx")
